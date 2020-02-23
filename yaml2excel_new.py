import tablib
import yaml
import click
import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import colors, Color, PatternFill, Font, Border, Side


greyfill = PatternFill(start_color="d4d0c8", end_color="d4d0c8", fill_type="solid")
redfill = PatternFill(start_color="f8cbad", end_color="f8cbad", fill_type="solid")
bluefill = PatternFill(start_color="79c5ca", end_color="fffffd", fill_type="solid")
greenfill = PatternFill(start_color="29e32e", end_color="29e32e", fill_type="solid")
border = Border(left=Side(style='thin',color='000000'), right=Side(style='thin',color='000000'), top=Side(style='thin',color='000000'), bottom=Side(style='thin',color='000000'))



def create_scan_code_excel(conf_path, entry_code, order_no, size_label):
    # 使用openpyxl 创建excel表格
    wb = Workbook()
    ws = wb.active
    dv = DataValidation(type='list', formula1='"1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24"', allow_blank=False)
    dv.prompt = "选择颜色"
    dv.promptTitle = 'List Selection'
    ws.add_data_validation(dv)
    # 设置表头
    entry_code_title = ws["A1"]
    entry_code_title.value = "entry_code"
    entry_code_title.fill = greyfill
    entry_code_title.border = border

    order_no_title = ws["B1"]
    order_no_title.value = "order_no"
    order_no_title.fill = greyfill
    order_no_title.border = border

    style_id_title = ws["C1"]
    style_id_title.value = "style_id"
    style_id_title.fill = greyfill
    style_id_title.border = border

    batch_no_title = ws["D1"]
    batch_no_title.value = "batch_no"
    batch_no_title.fill = greyfill
    batch_no_title.border = border

    color_no_title = ws["E1"]
    color_no_title.value = "color_no"
    color_no_title.fill = greyfill
    color_no_title.border = border

    size_label_title = ws["F1"]
    size_label_title.value = "size_label"
    size_label_title.fill = greyfill
    size_label_title.border = border

    quantity_title = ws["G1"]
    quantity_title.value = "quantity"
    quantity_title.fill = greyfill
    quantity_title.border = border

    row = ["A", "B", "C", "D", "E", "F", "G"]
    with open(conf_path, encoding='utf-8') as f:
        content = yaml.safe_load(f.read())
    
    info = content.get('info')
    style_id = info.get('style_id')

    # 组织数据
    data_list = [[entry_code, order_no, style_id, "", "24", size_label[i], "1"] for i in range(len(size_label))]
    # 填充数据
    for data in data_list:
        current_row = ws.max_row
        for i, v in enumerate(row):
            if v == 'E':
                dv.add(ws[v + str(current_row + 1)])
            else:
                ws[v + str(current_row + 1)].value = data[i]
            ws[v + str(current_row + 1)].fill = redfill
            ws[v + str(current_row + 1)].border = border
            
    wb.save("1.xxxx.xlsx")


def create_size_configure_excel(conf_path, order, size):
    headers = ('order_no', 'size_label', 'num_id', 'name_index', 'fold', 'unit', 'check', 'std_val', 'tol_lower', 'tol_upper', 'offset')
    with open(conf_path, encoding='utf-8') as f:
        content = yaml.safe_load(f.read())
    result = []
    steps = content.get('steps')
    for s in size:
        for i in steps:
            std_val_map = {k: v.get("val") for k, v in i.get("std_val").items()} if i.get("std_val") else {}
            std_val_list = list(std_val_map.keys())
            measure_list = list(i.get('measure').keys()) if i.get('measure') else []
            diff_list = list(i.get('diff').keys()) if i.get('diff') else []
            num_id_list = set(std_val_list + measure_list + diff_list)
            dataset = [(order, s, i, '01', '', 'Inch', '', std_val_map.get(i), '', '', '') for i in num_id_list]
            result += dataset

    data = tablib.Dataset(*result, headers=headers, title="sheet1")
    open("2.xxxx.xlsx", "wb").write(data.xlsx)
    
    # 继续对表格进行处理
    excel = load_workbook("2.xxxx.xlsx")
    table = excel.get_sheet_by_name("sheet1")
    row = table.max_row
    dv = DataValidation(type='list', formula1='"Inch", "cm"', allow_blank=False)
    dv.prompt = "选择单位"
    dv.promptTitle = 'List Selection'
    table.add_data_validation(dv)
    for i in range(2, row + 1):
        dv.add(table["F" + str(i)])

    # 添加单元格颜色
    title = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
    for i in range(2, row + 1):
        for t in title:
            table[t + "1"].fill = greyfill
            table[t + "1"].border = border
            if table["B" + str(i)].value in size and size.index(table["B" + str(i)].value) % 2 == 0:
                table[t + str(i)].fill = redfill
            elif table["B" + str(i)].value in size and size.index(table["B" + str(i)].value) % 2 == 1:
                table[t + str(i)].fill = bluefill
            table[t + str(i)].border = border
    excel.save("2.xxxx.xlsx")


def create_style_step_numid_excel(conf_path):
    headers = ('style_id', 'step', 'num_id')
    with open(conf_path, encoding='utf-8') as f:
        content = yaml.safe_load(f.read())
    result = []
    info = content.get('info')
    style_id = info.get('style_id')
    steps = content.get('steps')
    for i, v in enumerate(steps):
        step = v.get('index')
        std_val_list = list(v.get('std_val').keys()
                            ) if v.get('std_val') else []
        measure_list = list(v.get('measure').keys()
                            ) if v.get('measure') else []
        diff_list = list(v.get('diff').keys()) if v.get('diff') else []
        num_id_list = set(std_val_list + measure_list + diff_list)
        dataset = [(style_id, step, x) for x in list(num_id_list)]
        result += dataset
    data = tablib.Dataset(*result, headers=headers, title="sheet1")
    open("3.xxxx.xlsx", "wb").write(data.xlsx)

    # 继续对表格进行处理
    excel = load_workbook("3.xxxx.xlsx")
    table = excel.get_sheet_by_name("sheet1")
    row = table.max_row
    title = ["A", "B", "C"]
    for i in range(2, row + 1):
        for t in title:
            table[t + "1"].fill = greyfill
            table[t + "1"].border = border
            table[t + str(i)].fill = redfill
            table[t + str(i)].border = border
    excel.save("3.xxxx.xlsx")



def create_style_step_property_excel(conf_path):
    headers = ('style_id', 'step', 'step_display_name', 'thumbnail', 'kpshow', 'keypoint')
    with open(conf_path, encoding='utf-8') as f:
        content = yaml.safe_load(f.read())
    result = []
    info = content.get('info')
    style_id = info.get('style_id')
    steps = content.get('steps')
    for i in steps:
        step = i.get("index")
        display_name = i.get("display_name")
        thumbnail = i.get("thumbnail")
        kpshow = i.get("kpshow")
        keypoint = i.get("keypoint")
        dataset = [style_id, step, display_name, thumbnail, kpshow, keypoint]
        result.append(dataset)
    data = tablib.Dataset(*result, headers=headers, title="sheet1")
    open("4.xxxx.xlsx", "wb").write(data.xlsx)

    # 继续对表格进行处理
    excel = load_workbook("4.xxxx.xlsx")
    table = excel.get_sheet_by_name("sheet1")
    row = table.max_row
    title = ["A", "B", "C", "D", "E", "F"]
    for i in range(2, row + 1):
        for t in title:
            table[t + "1"].fill = greyfill
            table[t + "1"].border = border
            table[t + str(i)].fill = redfill
            table[t + str(i)].border = border
    excel.save("4.xxxx.xlsx")


def create_numid_naming_sheet_excel(conf_path):
    headers = ['num_id']
    for i in range(1, 51):
        i = "0" + str(i) if i < 10 else str(i)
        headers.append(i)
    with open(conf_path, encoding='utf-8') as f:
        content = yaml.safe_load(f.read())
    result = []
    std_val_list = []
    measure_list = []
    diff_list = []
    steps = content.get('steps')
    for v in steps:
        std_val = v.get('std_val')
        measure = v.get('measure')
        diff = v.get('diff')
        if std_val:
            std_val_list = [(k, v.get("name")) + ('',) * 49 for k, v in std_val.items()]
        if measure:
            measure_list = [(k, v.get("name")) + ('',) * 49 for k, v in measure.items()]
        if diff:
            diff_list = [(k, v.get("name")) + ('',) * 49 for k, v in diff.items()]
        num_id_list = std_val_list + measure_list + diff_list
        for i in num_id_list:
            if i not in result:
                result += [i]
    data = tablib.Dataset(*result, headers=headers, title="sheet1")
    open("6.xxxx.xlsx", "wb").write(data.xlsx)

    # 继续对表格进行处理
    excel = load_workbook("6.xxxx.xlsx")
    table = excel.get_sheet_by_name("sheet1")
    row = table.max_row
    title = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", 
            "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF",
            "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY"]
    for i in range(2, row + 1):
        for t in title:
            table[t + "1"].fill = greyfill
            table[t + "1"].border = border
            table[t + str(i)].fill = bluefill
            table[t + str(i)].border = border
    excel.save("6.xxxx.xlsx")


def create_style_id_naming_sheet_excel(conf_path):
    headers = ('style_id', 'style_name_en')
    with open(conf_path, encoding='utf-8') as f:
        content = yaml.safe_load(f.read())
    info = content.get('info')
    style_id = info.get('style_id')
    style_name = info.get('style_name')
    data = tablib.Dataset((style_id, style_name), headers=headers, title="sheet1")
    open("8.xxxx.xlsx", "wb").write(data.xlsx)

    # 继续对表格进行处理
    excel = load_workbook("8.xxxx.xlsx")
    table = excel.get_sheet_by_name("sheet1")
    row = table.max_row
    title = ["A", "B"]
    for i in range(2, row + 1):
        for t in title:
            table[t + "1"].fill = greyfill
            table[t + "1"].border = border
            table[t + str(i)].fill = bluefill
            table[t + str(i)].border = border
    excel.save("8.xxxx.xlsx")


@click.command()
@click.option("-p", help="需要转换的yaml配置文件的路径", required=True)
@click.option("-entry", help="entry_code", required=True)
@click.option("-order", help="order_no", required=True)
@click.option("-size", help="size_label list", required=True)
def main(p, entry, order, size):
    suffix = os.path.splitext(p)[1]
    if suffix != ".yaml":
        raise SyntaxError("只能转换yaml文件！")
    if os.path.isfile(p) and os.path.exists(p):
        size = size.split(",")
        create_scan_code_excel(p, entry, order, size)
        create_size_configure_excel(p, order, size)
        create_style_step_numid_excel(p)
        create_style_step_property_excel(p)
        create_numid_naming_sheet_excel(p)
        create_style_id_naming_sheet_excel(p)
    else:
        raise FileNotFoundError("文件不存在或者不是单文件！")


if __name__ == "__main__":
    main()