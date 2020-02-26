import time
import tablib
import yaml
import click
import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import colors, Color, PatternFill, Font, Border, Side


greyfill = PatternFill(start_color="d4d0c8",
                       end_color="d4d0c8", fill_type="solid")
redfill = PatternFill(start_color="f8cbad",
                      end_color="f8cbad", fill_type="solid")
bluefill = PatternFill(start_color="79c5ca",
                       end_color="fffffd", fill_type="solid")
greenfill = PatternFill(start_color="29e32e",
                        end_color="29e32e", fill_type="solid")
border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))


def init_data(conf_path):
    content = get_yaml_content(conf_path)
    info = content.get('info')
    style_id = info.get('style_id')
    style_name = info.get('style_name')
    steps = content.get('steps')
    result = get_step_data(steps)

    return {"style_id": style_id, "style_name": style_name, "step_list": result}


def get_step_data(steps):
    result = []
    for step in steps:
        step_index = step.get("index")
        display_name = step.get("display_name")
        thumbnail = step.get("thumbnail")
        model_data = get_model_data(step)
        data = {"step_index": step_index, "display_name": display_name,
                "thumbnail": thumbnail, **model_data}
        result.append(data)

    return result


def get_model_data(step):
    std_val_map = {}
    std_val_list = []
    std_name_map = {}
    diff_list = []
    diff_name_map = {}
    measure_list = []
    measure_name_map = {}
    for m in step.get("models"):
        if m.get("type") == "std_val":
            std_val_map.update({m.get("uid"): m.get("val")})
            std_val_list.append(m.get("uid"))
            std_name_map.update({m.get("uid"): m.get("name")})
        if m.get("type") == "diff":
            diff_list.append(m.get("uid"))
            diff_name_map.update({m.get("uid"): m.get("name")})
        if m.get("measure"):
            measure_list.append(m.get("uid"))
            measure_name_map.update({m.get("uid"): m.get("name")})
    num_id_list = set(std_val_list + measure_list + diff_list)
    name_dict = {**std_name_map, **measure_name_map, **diff_name_map}

    return {"std_val_map": std_val_map, "std_val_list": std_val_list, "std_name_map": std_name_map,
            "diff_list": diff_list, "diff_name_map": diff_name_map, "measure_list": measure_list, "measure_name_map": measure_name_map,
            "num_id_list": num_id_list, "name_dict": name_dict}


def create_scan_code_excel(data_param, entry_code, order_no, size_label):
    headers = ("entry_code", "order_no", "style_id",
               "batch_no", "color_no", "size_label", "quantity")
    filename = "1.扫码表scan_code.xlsx"
    style_id = data_param.get('style_id')
    result = []
    for i in size_label:
        dataset = [entry_code, order_no, style_id, "", "24", i, "1"]
        result.append(dataset)
    data = tablib.Dataset(*result, headers=headers, title="sheet1")
    open(filename, "wb").write(data.xlsx)

    # 继续对表格进行处理
    excel, table, row = get_table_object(filename)
    formula1 = '"1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24"'
    prompt = "选择颜色"
    prompt_title = 'List Selection'
    column = "E"
    add_excel_choice(formula1, prompt, prompt_title, table, column, row)
    # 添加单元格颜色
    title = ["A", "B", "C", "D", "E", "F", "G"]
    add_excel_bgcolor_border(excel, table, row, filename, title)


def create_size_configure_excel(data_param, order, size):
    headers = ('order_no', 'size_label', 'num_id', 'name_index', 'fold',
               'unit', 'check', 'std_val', 'tol_lower', 'tol_upper', 'offset', 'name')
    filename = "2.尺寸表size_configure.xlsx"
    result = []
    step_list = data_param.get('step_list')
    for s in size:
        for step_param in step_list:
            std_val_map = step_param.get('std_val_map')
            num_id_list = step_param.get('num_id_list')
            name_dict = step_param.get('name_dict')

            dataset = [(order, s, i, '01', '1', 'Inch', 'Y', std_val_map.get(
                i), '', '', '0' if i in diff_list else '', name_dict.get(i)) for i in num_id_list]
            result += dataset

    data = tablib.Dataset(*result, headers=headers, title="sheet1")
    open(filename, "wb").write(data.xlsx)

    # 继续对表格进行处理
    excel, table, row = get_table_object(filename)
    # 增加尺寸的选择
    formula1_1 = '"Inch, cm"'
    prompt_1 = "选择单位"
    prompt_title = 'List Selection'
    column_1 = "F"
    add_excel_choice(formula1_1, prompt_1, prompt_title, table, column_1, row)

    # 增加check的选择
    formula1_2 = '"Y, N"'
    prompt_2 = "选择Y/N"
    column_2 = "G"
    add_excel_choice(formula1_2, prompt_2, prompt_title, table, column_2, row)

    # 添加单元格颜色
    title = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
    for i in range(2, row + 1):
        for t in title:
            table[t + "1"].fill = greyfill
            table[t + "1"].border = border
            if table["B" + str(i)].value in size and size.index(table["B" + str(i)].value) % 2 == 0:
                table[t + str(i)].fill = redfill
            elif table["B" + str(i)].value in size and size.index(table["B" + str(i)].value) % 2 == 1:
                table[t + str(i)].fill = bluefill
            table[t + str(i)].border = border
    excel.save(filename)


def create_style_step_numid_excel(data_param):
    headers = ('style_id', 'step', 'num_id')
    filename = "3.款式表style_step_num_id.xlsx"
    style_id = data_param.get('style_id')
    step_list = data_param.get('step_list')

    result = []
    for step_param in step_list:
        step_index = step_param.get('step_index')
        num_id_list = step_param.get('num_id_list')
        dataset = [(style_id, step_index, x) for x in list(num_id_list)]
        result += dataset
    data = tablib.Dataset(*result, headers=headers, title="sheet1")
    open(filename, "wb").write(data.xlsx)

    # 继续对表格进行处理
    title = ["A", "B", "C"]
    excel, table, row = get_table_object(filename)
    add_excel_bgcolor_border(excel, table, row, filename, title)


def create_style_step_property_excel(data_param):
    headers = ('style_id', 'step', 'step_display_name', 'thumbnail')
    filename = "4.款式配置表style_step_property.xlsx"
    style_id = data_param.get('style_id')
    step_list = data_param.get('step_list')
    result = []
    for step_param in step_list:
        step_index = step_param.get('step_index')
        display_name = step_param.get('display_name')
        thumbnail = step_param.get('thumbnail')
        dataset = [style_id, step_index, display_name, thumbnail]
        result.append(dataset)

    data = tablib.Dataset(*result, headers=headers, title="sheet1")
    open(filename, "wb").write(data.xlsx)

    # 继续对表格进行处理
    title = ["A", "B", "C", "D"]
    excel, table, row = get_table_object(filename)
    add_excel_bgcolor_border(excel, table, row, filename, title)


def create_numid_naming_sheet_excel(data_param):
    headers = ['num_id']
    for i in range(1, 51):
        i = "0" + str(i) if i < 10 else str(i)
        headers.append(i)
    filename = "6.客制测量项表num_id_naming_sheet.xlsx"
    result = []
    step_list = data_param.get('step_list')
    for step_param in step_list:
        std_name_map = step_param.get('std_name_map')
        measure_name_map = step_param.get('measure_name_map')
        diff_name_map = step_param.get('diff_name_map')

        std_val_list = [(k, v) + ('',) * 49 for k, v in std_name_map.items()]
        measure_list = [(k, v) + ('',) * 49 for k,
                        v in measure_name_map.items()]
        diff_list = [(k, v) + ('',) * 49 for k, v in diff_name_map.items()]

        num_id_list = std_val_list + measure_list + diff_list
        result += num_id_list

    data = tablib.Dataset(*result, headers=headers, title="sheet1")
    open(filename, "wb").write(data.xlsx)

    # 继续对表格进行处理
    title = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P",
             "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF",
             "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX",
             "AY"]
    excel, table, row = get_table_object(filename)
    add_excel_bgcolor_border(excel, table, row, filename, title)


def create_style_id_naming_sheet_excel(data_param):
    headers = ('style_id', 'style_name_en')
    filename = "8.款式编码表style_id_naming_sheet.xlsx"
    style_id = data_param.get('style_id')
    style_name = data_param.get('style_name')
    data = tablib.Dataset((style_id, style_name),
                          headers=headers, title="sheet1")
    open(filename, "wb").write(data.xlsx)
    # 继续对表格进行处理
    title = ["A", "B"]
    excel, table, row = get_table_object(filename)
    add_excel_bgcolor_border(excel, table, row, filename, title)


def get_yaml_content(conf_path):
    content = None
    with open(conf_path, encoding='utf-8') as f:
        content = yaml.safe_load(f.read())
    return content


def add_excel_bgcolor_border(excel, table, row, filename, title):
    for i in range(2, row + 1):
        for t in title:
            table[t + "1"].fill = greyfill
            table[t + "1"].border = border
            table[t + str(i)].fill = bluefill
            table[t + str(i)].border = border
    excel.save(filename)


def get_table_object(filename):
    excel = load_workbook(filename)
    table = excel.get_sheet_by_name("sheet1")
    row = table.max_row
    return excel, table, row


def add_excel_choice(formula1, prompt, prompt_title, table, column, row):
    dv = DataValidation(type='list', formula1=formula1, allow_blank=False)
    dv.prompt = prompt
    dv.promptTitle = prompt_title
    table.add_data_validation(dv)
    for i in range(2, row + 1):
        dv.add(table[column + str(i)])


@click.command()
@click.option("-p", help="需要转换的yaml配置文件的路径", required=True)
@click.option("-entry", help="entry_code", required=True)
@click.option("-order", help="order_no", required=True)
@click.option("-size", help="size_label list", required=True)
def main(p, entry, order, size):
    a = time.time()
    suffix = os.path.splitext(p)[1]
    if suffix != ".yaml":
        raise SyntaxError("只能转换yaml文件！")
    if os.path.isfile(p) and os.path.exists(p):
        data_param = init_data(p)
        size = size.split(",")
        create_scan_code_excel(data_param, entry, order, size)
        create_size_configure_excel(data_param, order, size)
        create_style_step_numid_excel(data_param)
        create_style_step_property_excel(data_param)
        create_numid_naming_sheet_excel(data_param)
        create_style_id_naming_sheet_excel(data_param)

        print("时间: {}".format(time.time() - a))
    else:
        raise FileNotFoundError("文件不存在或者不是单文件！")


if __name__ == "__main__":
    main()
